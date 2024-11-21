import win32com.client
import openpyxl
import os
import re
import shutil
import datetime
import pythoncom
# Constants for Excel orientation
xlPortrait = 1
xlLandscape = 2

current_dir = os.getcwd()
horizontal_jump_card_path = os.path.join(current_dir, 'static','cards','horizontal_jump_card.html')
distance_card_path = os.path.join(current_dir, 'static','cards','distance_card.html')
height_card_path = os.path.join( current_dir, 'static', 'cards', 'height_card.html')
lff_dir = os.path.join(current_dir, 'templates', 'cards')
pdf_file_path = os.path.join(current_dir, 'output.pdf')
output_pdfs_path = os.path.join(current_dir, 'static', 'generated_files')
lff_dir = current_dir

#Add Meeting Name, Venue Name and Event Name
meeting_name_pattern = re.compile(r'<m>')
venue_name_pattern = re.compile(r'<v>')
event_name_pattern = re.compile(r'<e>')
created_date_pattern = re.compile(r'<c>')
height_metres_pattern = re.compile(r'<metres>')
row_pattern = re.compile(r'&nbsp;&nbsp;')
distance_lffs_pattern = re.compile(r'\b(Discus|Shot|Hammer|Throw|Javelin)\b', re.IGNORECASE)

#!!!Alternative with Excel
def export_fieldcards_to_excel_then_pdf(meeting_name,venue_name,lff_dir,lff_files, pdf_file_path):
    pythoncom.CoInitialize() #Initialize COM Library
    
    
    excel_file_path = os.path.join(current_dir, 'static','cards','Fieldcard Templates.xlsx')
    distance_lffs_pattern = re.compile(r'\b(Discus|Shot|Hammer|Throw|Javelin)\b', re.IGNORECASE)
    worksheet_print_no = 0 #? Worksheet number in the book for printing
    
    
    for filename in lff_files:
        file_path = os.path.join(lff_dir, filename)
        with open(file_path, 'r') as file:
            lines = file.readlines()
            event_line = lines[0].strip().split(',')
            event_name = event_line[3]
            
            
        # Create a temporary copy of the workbook
        temp_workbook_path = os.path.join(os.path.dirname(excel_file_path), f"temp_{os.path.basename(excel_file_path)}")
        shutil.copy2(excel_file_path, temp_workbook_path)
        
        # Open the Excel file
        workbook = openpyxl.load_workbook(temp_workbook_path)
        
        if "Long" in event_name or "Triple" in event_name: #? Horizontal Jump FieldCards
            worksheet_print_no = 1 #? Horizontal Cards are at no 1
            #Get the Horzontal Jump Card sheet
            sheet = workbook.worksheets[0]
        
            sheet['C3'] = meeting_name
            sheet['H3'] = venue_name
            sheet['M2'] = event_name
            #Set the created date to R3
            sheet['R3'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            blank_rows = []
            for row_index, line in enumerate(lines):
                if row_index == 0:
                    continue
                if not line or line == " " or line == "\n":
                    continue
                
                line_parts = line.strip().split(',')
                
                bib_no = line_parts[1]
                name = line_parts[5] + ' ' + line_parts[4]
                team = line_parts[6]
                
                trial_data_list = line_parts[7:]
                
                if name == "" or name == " " and line_parts[7] == "DNS":
                    blank_rows.append(row_index)
                    continue
                
                if blank_rows:
                    row_index = blank_rows.pop(0)
                    #print("New Index", row_index)
                    
                if name != " " and line_parts[7] == "DNS":
                    sheet.cell(row=start_row + row_index, column=2, value=bib_no)
                    sheet.cell(row=start_row + row_index, column=3, value=name)
                    sheet.cell(row=start_row + row_index, column=4, value=team)
                    sheet.cell(row=start_row + row_index, column=20, value="DNS")
                    continue
                    
                
                # Define starting column index (E corresponds to column 5 in 1-based index)
                first_col_section = 5 # Column E
                second_col_section = 13 # Column M
                start_row = 6
                
                # Separate the first five elements and the remaining elements
                first_section = trial_data_list[:6]
                second_section = trial_data_list[6:]

                # Function to convert to float and ignore strings
                def to_float(value):
                    try:
                        return float(value)
                    except ValueError:
                        return None

                # Collect alternating values and convert to float
                first_section_values = [to_float(first_section[i]) for i in range(0, len(first_section), 2) if to_float(first_section[i]) is not None]
                second_section_values = [to_float(second_section[i]) for i in range(0, len(second_section), 2) if to_float(second_section[i]) is not None]

                print(first_section_values,second_section_values)

                # Calculate the highest values
                max_first_section = max(first_section_values) if first_section_values else 0
                max_second_section = max(second_section_values) if second_section_values else 0

                #Write the Bib, Name and Team
                sheet.cell(row=start_row + row_index, column=2, value=bib_no)
                sheet.cell(row=start_row + row_index, column=3, value=name)
                sheet.cell(row=start_row + row_index, column=4, value=team)
                
                # Write the first five elements starting from column E
                for i in range(6):
                    #print(trial_data_list[i])
                    #If the value is a string then write "X"
                    if trial_data_list[i] == "F":
                        sheet.cell(row=start_row + row_index, column=first_col_section + i, value="X")
                    else:
                        sheet.cell(row=start_row + row_index, column=first_col_section + i, value=trial_data_list[i])
                    
                #Write the Best of 3 and the position
                sheet.cell(row=start_row + row_index, column=first_col_section + 6, value= "" if max_first_section == 0 else max_first_section)
                sheet.cell(row=start_row + row_index, column=first_col_section + 7, value= row_index)
                
                # Skip two columns and continue writing the rest of the data
                for i in range(len(second_section)):
                    #If the value is a string then write "X"
                    if trial_data_list[i] == "F":
                        sheet.cell(row=start_row + row_index, column=second_col_section + i, value="X")
                    else:
                        sheet.cell(row=start_row + row_index, column=second_col_section + i, value=trial_data_list[i])
                    
                #Write the Best of All and the final Position
                sheet.cell(row=start_row + row_index, column=second_col_section + 6, value= "" if max(max_first_section, max_second_section) == 0 else max(max_first_section, max_second_section))
                sheet.cell(row=start_row + row_index, column=second_col_section + 7, value= row_index) 
                
        elif bool(distance_lffs_pattern.search(event_name)): #? Distance LFFs
            worksheet_print_no = 2 #? Distance Cards are at no 2
            
            #Get the Horzontal Jump Card sheet
            sheet = workbook.worksheets[1]
        
            sheet['C3'] = meeting_name
            sheet['G3'] = venue_name
            sheet['J2'] = event_name
            #Set the created date to M3
            sheet['M3'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            blank_rows = []
            for row_index, line in enumerate(lines):
                if row_index == 0:
                    continue
                if not line or line == " " or line == "\n":
                    continue
                
                line_parts = line.strip().split(',')
                
                bib_no = line_parts[1]
                name = line_parts[5] + ' ' + line_parts[4]
                team = line_parts[6]
                
                trial_data_list = line_parts[7:]
                
                if name == "" or name == " " and line_parts[7] == "DNS":
                    blank_rows.append(row_index)
                    continue
                
                if blank_rows:
                    row_index = blank_rows.pop(0)
                    #print("New Index", row_index)
                    
                if name != " " and line_parts[7] == "DNS":
                    sheet.cell(row=start_row + row_index, column=2, value=bib_no)
                    sheet.cell(row=start_row + row_index, column=3, value=name)
                    sheet.cell(row=start_row + row_index, column=4, value=team)
                    continue
                    
                
                # Define starting column index (E corresponds to column 5 in 1-based index)
                first_col_section = 5 # Column E
                second_col_section = 10 # Column J
                start_row = 6
                
                # Separate the first five elements and the remaining elements
                first_section = trial_data_list[:6]
                second_section = trial_data_list[6:]

                # Function to convert to float and ignore strings
                def to_float(value):
                    try:
                        return float(value)
                    except ValueError:
                        return None

                # Collect alternating values and convert to float
                first_section_values = [to_float(first_section[i]) for i in range(0, len(first_section), 2) if to_float(first_section[i]) is not None]
                second_section_values = [to_float(second_section[i]) for i in range(0, len(second_section), 2) if to_float(second_section[i]) is not None]

                print(first_section_values,second_section_values)

                # Calculate the highest values
                max_first_section = max(first_section_values) if first_section_values else 0
                max_second_section = max(second_section_values) if second_section_values else 0

                #Write the Bib, Name and Team
                sheet.cell(row=start_row + row_index, column=2, value=bib_no)
                sheet.cell(row=start_row + row_index, column=3, value=name)
                sheet.cell(row=start_row + row_index, column=4, value=team)
                
                print(trial_data_list)
                print(max_first_section,max_second_section)
                
                # Write the first five elements starting from column E #? The wind values are ignored explaining the step_index in range
                column_index = 0
                for i in range(0,6,2):
                    #If the value is a string then write "X"
                    if trial_data_list[i] == "F":
                        sheet.cell(row=start_row + row_index, column=first_col_section + column_index , value="X")
                    else:
                        sheet.cell(row=start_row + row_index, column=first_col_section + column_index , value=trial_data_list[i])
                    
                    if column_index == 3:
                        break
                    
                    column_index += 1
                #Write the Best of 3 and the position
                sheet.cell(row=start_row + row_index, column=first_col_section + 3, value= "" if max_first_section == 0 else max_first_section)
                sheet.cell(row=start_row + row_index, column=first_col_section + 4, value= row_index)
                
                # Skip two columns and continue writing the rest of the data #? The wind values are ignored explaining the step_index in range
                column_index = 0
                for i in range(6,len(trial_data_list),2):
                    #If the value is a string then write "X"
                    if trial_data_list[i] == "F":
                        sheet.cell(row=start_row + row_index, column=second_col_section + column_index, value="X")
                    else:
                        sheet.cell(row=start_row + row_index, column=second_col_section + column_index, value=trial_data_list[i])
                        
                    if column_index == 3:
                        break
                    
                    column_index += 1  
                    
                #Write the Best of All and the final Position
                #print("" if max(max_first_section, max_second_section) == 0 ))
                sheet.cell(row=start_row + row_index, column=second_col_section + 3, value= "" if max(max_first_section, max_second_section) == 0 else max(max_first_section, max_second_section))
                sheet.cell(row=start_row + row_index, column=second_col_section + 4, value= row_index) 
        # Save the temp workbook
        workbook.save(temp_workbook_path)
        workbook.close()
        
        # Open Excel application
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False  # Make Excel application invisible

        # Open the workbook in Excel
        workbook_obj = excel_app.Workbooks.Open(temp_workbook_path)
        # Set print settings
        sheet_obj = workbook_obj.Worksheets(worksheet_print_no)
        sheet_obj.PageSetup.Orientation = xlLandscape  # Set orientation to landscape
        #sheet_obj.PageSetup.Zoom = True  # Enable zoom control
        sheet_obj.PageSetup.FitToPagesTall = 1  # Don't fit rows to pages
        sheet_obj.PageSetup.FitToPagesWide = 1  # Fit all columns on one page
        sheet_obj.PageSetup.PrintArea = sheet_obj.UsedRange.Address  # Set print area to used range

        # Adjust scaling and margins
        sheet_obj.PageSetup.Zoom = 75  # Set zoom to 100%
        sheet_obj.PageSetup.TopMargin = excel_app.InchesToPoints(0.25)  # Top margin 0.25 inches
        sheet_obj.PageSetup.BottomMargin = excel_app.InchesToPoints(0.25)  # Bottom margin 0.25 inches
        sheet_obj.PageSetup.LeftMargin = excel_app.InchesToPoints(0.25)  # Left margin 0.25 inches
        sheet_obj.PageSetup.RightMargin = excel_app.InchesToPoints(0.25)  # Right margin 0.25 inches


        # Export to PDF
        pdf_file = pdf_file_path
        sheet_obj.ExportAsFixedFormat(0, os.path.join(current_dir, f'{filename}.pdf'))

        # Close Excel application
        workbook_obj.Close(SaveChanges=False)
        excel_app.Quit()
        
        os.remove(temp_workbook_path)


def export_fieldcards_to_pdf(meeting_name,venue_name,lff_dir,lff_files, pdf_file_path):
    
    new_html_files = [] #*Used to store the paths for the new html files
    
    for filename in lff_files:
        file_path = os.path.join(lff_dir, filename)
        with open(file_path, 'r') as file:
            lines = file.readlines()
            #data = file.read()
            event_line = lines[0].strip().split(',')
            event_name = event_line[3]
        
        # Split the data into rows
        rows = [row.strip('\n').split(',') for row in lines[1:]]
            #print(rows)
        # Modify the HTML content
        modified_html = ''
        current_row = 0

        modified_rows = []
        blank_rows = []
        
        if bool(distance_lffs_pattern.search(event_name)): #? Distance LFFs
            # Read the HTML file for the Distance Cards
            with open(distance_card_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
                
            
            for row_index, line_parts in enumerate(rows):
                row_list = []
                if not line_parts or line_parts == " " or line_parts == "\n":
                    continue
                
                bib_no = line_parts[1]
                name = line_parts[5] + ' ' + line_parts[4]
                team = line_parts[6]
                
                row_list.append(bib_no)
                row_list.append(name)
                row_list.append(team)
                
                trial_data_list = line_parts[7:]
                
                if name == "" or name == " " and line_parts[7] == "DNS":
                    blank_rows.append(row_index)
                    continue
                            
                if blank_rows:
                    row_index = blank_rows.pop(0)
                
                if name != " " and line_parts[7] == "DNS":
                    row_list.extend(['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'DNS'])
                    continue
                
                #trial_data_list = ['X' if value == 'F' else value for value in trial_data_list]
                
                # Separate the first five elements and the remaining elements
                first_section = trial_data_list[:6]
                second_section = trial_data_list[6:]
                # Function to convert to float and ignore strings
                def to_float(value):
                    try:
                        return float(value)
                    except ValueError:
                        if value == "F":
                            return 0.0 #? Marks F values 
                        elif value == '-':
                            return 0.0001 #? Marks - value from the lff files
                # Collect alternating values and convert to float
                first_section_values = [to_float(first_section[i]) for i in range(0, len(first_section), 2) if to_float(first_section[i]) is not None]
                second_section_values = [to_float(second_section[i]) for i in range(0, len(second_section), 2) if to_float(second_section[i]) is not None]
                #print(first_section_values,second_section_values)
                # Calculate the highest values
                max_first_section = max(first_section_values) if first_section_values else 0
                max_second_section = max(second_section_values) if second_section_values else 0

                
                total_max_value = "" if max(max_first_section, max_second_section) in (0, 0.0, 0.0001 ) else max(max_first_section, max_second_section)
                row_list.extend(first_section_values)
                row_list.append("" if (max_first_section == 0 or max_first_section == 0.0001) else max_first_section) #? Best of 3 Trials value
                row_list.append("" if (max_first_section == 0 or max_first_section == 0.0001) else row_index + 1)   #? Position after 3 value
                
                row_list.extend(second_section_values)
                row_list.append(total_max_value) #? Best of All Trials value
                row_list.append("" if total_max_value == "" else row_index + 1) #? Final Position value
                
                row_list = ['X' if value == 0.0 else value for value in row_list] #? Replace the marked 'F' values with 0.0 to 'X'
                row_list = ['-' if value == 0.0001 else value for value in row_list] #? Replace the marked '-' values with 0.0001 back to '-'
                row_list = [format(value, ".2f") if type(value) == float else value for value in row_list] #? Format all float values to two decimals places

                modified_rows.append(row_list)
            #print(modified_rows)
            for line in html_content.split('\n'):
                match = meeting_name_pattern.search(line)
                if match:
                    line = meeting_name_pattern.sub(meeting_name,line)
                
                match = venue_name_pattern.search(line)
                if match:
                    line = venue_name_pattern.sub(venue_name, line)
                
                match = event_name_pattern.search(line)
                if match:
                    line = event_name_pattern.sub(event_name, line)
                    
                match = created_date_pattern.search(line)
                if match:
                    date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    line = created_date_pattern.sub(date, line)
                    
                match = row_pattern.search(line)
                if match:
                    if current_row < len(modified_rows):
                        #print(modified_rows[current_row])
                        line = row_pattern.sub(str(modified_rows[current_row].pop(0)), line)
                        if not modified_rows[current_row]:
                            current_row += 1
                modified_html += line + '\n'

            # # Write the modified HTML content to a new file
            # with open('templates/cards/output.html', 'w', encoding='utf-8') as file:
            #     file.write(modified_html)
            #     file.close()
            
        elif "Long" in event_name or "Triple" in event_name: #? Horizontal Jump FieldCards
            with open(horizontal_jump_card_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            
            for row_index, line_parts in enumerate(rows):
                row_list = []
                if not line_parts or line_parts == " " or line_parts == "\n":
                    continue
                
                bib_no = line_parts[1]
                name = line_parts[5] + ' ' + line_parts[4]
                team = line_parts[6]
                
                row_list.append(bib_no)
                row_list.append(name)
                row_list.append(team)
                
                trial_data_list = line_parts[7:]
                
                if name == "" or name == " " and line_parts[7] == "DNS":
                    blank_rows.append(row_index)
                    continue
                            
                if blank_rows:
                    row_index = blank_rows.pop(0)
                
                if name != " " and line_parts[7] == "DNS":
                    row_list.extend(['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'DNS'])
                    modified_rows.append(row_list)
                    continue

                trial_data_list = ['X' if value == 'F' else value for value in trial_data_list]
                
                # Separate the first five elements and the remaining elements
                first_section = trial_data_list[:6]
                second_section = trial_data_list[6:]
                # Function to convert to float and ignore strings
                def to_float(value):
                    try:
                        return float(value)
                    except ValueError:
                        if value == "X":
                            return 0.0 #? Marks F values 
                        elif value == '-':
                            return 0.0001 #? Marks - value from the lff files
                # Collect alternating values and convert to float
                first_section_values = [to_float(first_section[i]) for i in range(0, len(first_section), 2) if to_float(first_section[i]) is not None]
                second_section_values = [to_float(second_section[i]) for i in range(0, len(second_section), 2) if to_float(second_section[i]) is not None]
                #print(first_section_values,second_section_values)
                # Calculate the highest values
                max_first_section = max(first_section_values) if first_section_values else 0
                max_second_section = max(second_section_values) if second_section_values else 0

                
                total_max_value = "" if max(max_first_section, max_second_section) in (0, 0.0, 0.0001 ) else max(max_first_section, max_second_section)
                row_list.extend(first_section)
                row_list.append("" if (max_first_section == 0 or max_first_section == 0.0001) else max_first_section) #? Best of 3 Trials value
                row_list.append("" if (max_first_section == 0 or max_first_section == 0.0001) else row_index + 1)   #? Position after 3 value
                
                row_list.extend(second_section)
                row_list.append(total_max_value) #? Best of All Trials value
                row_list.append("" if total_max_value == "" else row_index + 1) #? Final Position value
                
                #row_list = ['X' if value == 0.0 else value for value in row_list] #? Replace the marked 'F' values with 0.0 to 'X'
                #row_list = ['-' if value == 0.0001 else value for value in row_list] #? Replace the marked '-' values with 0.0001 back to '-'
                row_list = [format(value, ".2f") if type(value) == float else value for value in row_list] #? Format all float values to two decimals places

                modified_rows.append(row_list)
            #print(modified_rows)
            for line in html_content.split('\n'):
                match = meeting_name_pattern.search(line)
                if match:
                    line = meeting_name_pattern.sub(meeting_name,line)
                
                match = venue_name_pattern.search(line)
                if match:
                    line = venue_name_pattern.sub(venue_name, line)
                
                match = event_name_pattern.search(line)
                if match:
                    line = event_name_pattern.sub(event_name, line)
                    
                match = created_date_pattern.search(line)
                if match:
                    date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    line = created_date_pattern.sub(date, line)
                    
                match = row_pattern.search(line)
                if match:
                    if current_row < len(modified_rows):
                        #print(modified_rows[current_row])
                        line = row_pattern.sub(str(modified_rows[current_row].pop(0)), line)
                        if not modified_rows[current_row]:
                            current_row += 1
                modified_html += line + '\n'
                
        elif "High" in event_name: #? High Jump LFF Events
            with open(height_card_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            
            # ?Find the indices of 'SH' (Start Height) and 'EH' (End Height)
            start_index = event_line.index('SH') + 1  # Start after 'SH'
            end_index = event_line.index('EH')  # Stop at 'EH'

            # ?Extract the parts between 'SH' and 'EH' and convert to float (Will be the event heights)
            event_heights = [float(event_line[i]) for i in range(start_index, end_index)]
            height_column = 0
            
            positions_list = [line.split(",")[3] for line in lines[1:]] #? Ignore the event row at index 0
            print("Positions", positions_list)
            #? Mark the duplicate positions with = sign
            occurrences = {}
            new_positions_list = []
            # First pass: count the occurrences of each number
            for num in positions_list:
                if num in occurrences:
                    occurrences[num] += 1
                else:
                    occurrences[num] = 1
                
            # Second pass: build the new list with '=' sign for duplicates
            for num in positions_list:
                if occurrences[num] > 1:
                    new_positions_list.append(f"{num}=")
                else:
                    new_positions_list.append(num)
            
            for row_index, line_parts in enumerate(rows):
                row_list = []
                if not line_parts or line_parts == " " or line_parts == "\n":
                    continue
                
                bib_no = line_parts[1]
                name = line_parts[5] + ' ' + line_parts[4]
                team = line_parts[6]
                
                row_list.append(bib_no)
                row_list.append(name)
                row_list.append(team)
                
                trial_data_list = line_parts[7:]
                
                #?Ensure that the list length is always equal to 12
                trial_data_list = trial_data_list + ['' for _ in range(12 - len(trial_data_list))] if len(trial_data_list) < 12 else trial_data_list
                
                if name == "" or name == " " and line_parts[7] == "DNS":
                    blank_rows.append(row_index)
                    continue
                            
                if blank_rows:
                    row_index = blank_rows.pop(0)
                
                if name != " " and line_parts[7] == "DNS":
                    row_list.extend(['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
                    modified_rows.append(row_list)
                    continue

                #trial_data_list = ['X' if value == 'F' else value for value in trial_data_list]
                
                # Initialize the variable to store the index of the last occurrence of 'o'
                last_o_index = -1
                retired_mark = False
                            
                # Iterate over the parts to find the last 'o'
                for l_index, part in enumerate(trial_data_list):
                    if 'o' in part or 'O' in part:
                        last_o_index = l_index  # Update the index where 'o' is found
                    
                    if 'r' in part or 'R' in part:
                        retired_mark = True
                        
                x_count = 0
                o_count = 0
                if last_o_index != -1:    
                    for i in range(last_o_index + 1):
                        x_count += trial_data_list[i].count('X')
                        o_count += trial_data_list[i].count('O')
                
                
                best_height_val =  "" if trial_data_list[0] == "" else ("" if last_o_index == -1 else format(event_heights[last_o_index], ".2f"))
                trials_at_best_height_val = "" if last_o_index == -1 else trial_data_list[last_o_index].index("O") + 1
                total_failures_val = "" if retired_mark else x_count
                total_trials_val = "" if retired_mark else x_count + o_count
                final_position = "" if retired_mark else new_positions_list[row_index]
                
                row_list.extend(trial_data_list)
                row_list.append(best_height_val)
                row_list.append(trials_at_best_height_val)
                row_list.append(total_failures_val)
                row_list.append(total_trials_val)
                row_list.append(final_position)
                                

                modified_rows.append(row_list)
            #print(modified_rows)
            for line in html_content.split('\n'):
                match = meeting_name_pattern.search(line)
                if match:
                    line = meeting_name_pattern.sub(meeting_name,line)
                
                match = venue_name_pattern.search(line)
                if match:
                    line = venue_name_pattern.sub(venue_name, line)
                
                match = event_name_pattern.search(line)
                if match:
                    line = event_name_pattern.sub(event_name, line)
                    
                match = created_date_pattern.search(line)
                if match:
                    date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    line = created_date_pattern.sub(date, line)
                
                match = height_metres_pattern.search(line)
                if match:
                    if height_column < len(event_heights):
                        #print(event_heights[height_column])
                        line = height_metres_pattern.sub(str(event_heights[height_column]), line)
                        #if not event_heights[height_column]:
                        height_column += 1
                            
                match = row_pattern.search(line)
                if match:
                    if current_row < len(modified_rows):
                        #print(modified_rows[current_row])
                        line = row_pattern.sub(str(modified_rows[current_row].pop(0)), line)
                        if not modified_rows[current_row]:
                            current_row += 1
                modified_html += line + '\n'

        # Write the modified HTML content to a new file
        #html_file_path = os.path.join(output_pdfs_path, f'{filename}.html')
        html_file_path = os.path.join(output_pdfs_path, f'{event_name}.html')
        with open(html_file_path, 'w', encoding='utf-8') as file:
            file.write(modified_html)
            file.close()
            
        new_html_files.append(html_file_path)
    return new_html_files

def export_evt_to_pdf(meeting_name,venue_name,lff_dir):
    new_html_files = [] #*Used to store the paths for the new html files
    
    lff_evt_path = os.path.join(lff_dir, "lynx.evt")
    
    lines = []
    if not os.path.exists(lff_evt_path):
        return f"No lynx.evt file was present in the lff directory. {lff_dir}"
        
    with open(lff_evt_path, 'r') as file:
        lines = file.readlines()    

    # Iterate over the lines and process the events
    i = 0
    while i < len(lines):
        # Identify an event row by checking the first and fourth elements
        modified_html = ""
        current_row = 0
        
        columns = lines[i].split(',')
        if columns[0].isdigit() and len(columns) > 3:
            event_name = columns[3]
            print(f"Event: {event_name}")
            
            # Collect all subsequent rows until the next event
            row_list = []
            i += 1
            while i < len(lines) and (not lines[i].split(',')[0].isdigit()):
                #print(lines[i])
                event_row = lines[i].split(",")
                
                bib_no = event_row[1]
                name = event_row[4] + " " + event_row[3]
                club = event_row[5]
                
                row_list.append([bib_no,name,club])
                i += 1
            
            if bool(distance_lffs_pattern.search(event_name)): #? Distance LFFs
                # Read the HTML file for the Distance Cards
                with open(distance_card_path, 'r', encoding='utf-8') as file:
                    html_content = file.read()
                #Append empty strings to each row in row list
                for row in row_list:
                    row.extend(["","","","","","","","","",""])
                print(row_list)
                
                for line in html_content.split('\n'):
                    match = meeting_name_pattern.search(line)
                    if match:
                        line = meeting_name_pattern.sub(meeting_name,line)
                    
                    match = venue_name_pattern.search(line)
                    if match:
                        line = venue_name_pattern.sub(venue_name, line)
                    
                    match = event_name_pattern.search(line)
                    if match:
                        line = event_name_pattern.sub(event_name, line)
                        
                    match = created_date_pattern.search(line)
                    if match:
                        date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        line = created_date_pattern.sub(date, line)
                        
                    match = row_pattern.search(line)
                    if match:
                        if current_row < len(row_list):
                            #print(modified_rows[current_row])
                            line = row_pattern.sub(str(row_list[current_row].pop(0)), line)
                            if not row_list[current_row]:
                                current_row += 1
                    modified_html += line + '\n'
                    
            elif "Long" in event_name or "Triple" in event_name: #? Horizontal Jump FieldCards
                with open(horizontal_jump_card_path, 'r', encoding='utf-8') as file:
                    html_content = file.read()
                
                #Append empty strings to each row in row list
                for row in row_list:
                    row.extend(["","","","","","","","","","","","","","","",""])
                print(row_list)
                
                for line in html_content.split('\n'):
                    match = meeting_name_pattern.search(line)
                    if match:
                        line = meeting_name_pattern.sub(meeting_name,line)
                    
                    match = venue_name_pattern.search(line)
                    if match:
                        line = venue_name_pattern.sub(venue_name, line)
                    
                    match = event_name_pattern.search(line)
                    if match:
                        line = event_name_pattern.sub(event_name, line)
                        
                    match = created_date_pattern.search(line)
                    if match:
                        date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        line = created_date_pattern.sub(date, line)
                        
                    match = row_pattern.search(line)
                    if match:
                        if current_row < len(row_list):
                            #print(modified_rows[current_row])
                            line = row_pattern.sub(str(row_list[current_row].pop(0)), line)
                            if not row_list[current_row]:
                                current_row += 1
                    modified_html += line + '\n'
            elif "High" in event_name: #? High Jump LFF Events
                with open(height_card_path, 'r', encoding='utf-8') as file:
                    html_content = file.read()
                    
                #Append empty strings to each row in row list
                for row in row_list:
                    row.extend(["","","","","","","","","","","","","","","","",""])
                print(row_list)
                
                for line in html_content.split('\n'):
                    match = meeting_name_pattern.search(line)
                    if match:
                        line = meeting_name_pattern.sub(meeting_name,line)
                    
                    match = venue_name_pattern.search(line)
                    if match:
                        line = venue_name_pattern.sub(venue_name, line)
                    
                    match = event_name_pattern.search(line)
                    if match:
                        line = event_name_pattern.sub(event_name, line)
                        
                    match = created_date_pattern.search(line)
                    if match:
                        date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        line = created_date_pattern.sub(date, line)
                    
                    # match = height_metres_pattern.search(line)
                    # if match:
                    #     if height_column < len(event_heights):
                    #         #print(event_heights[height_column])
                    #         line = height_metres_pattern.sub(str(event_heights[height_column]), line)
                    #         #if not event_heights[height_column]:
                    #         height_column += 1
                                
                    match = row_pattern.search(line)
                    if match:
                        if current_row < len(row_list):
                            #print(modified_rows[current_row])
                            line = row_pattern.sub(str(row_list[current_row].pop(0)), line)
                            if not row_list[current_row]:
                                current_row += 1
                    modified_html += line + '\n'

            html_file_path = os.path.join(output_pdfs_path, f'{event_name}.html')
            with open(html_file_path, 'w', encoding='utf-8') as file:
                file.write(modified_html)
                file.close()
                
            new_html_files.append(html_file_path)
        else:
            i += 1
    return new_html_files
    
    
#export_fieldcards_to_pdf("My Meeting","My Venue",lff_dir,["105-1-01.lff","108-1-01.lff","106-1-01.lff"],None)
#export_horizontal_jump_card_to_pdf("My Meeting","My Venue",lff_dir,["105-1-01.lff","109-1-01.lff"],pdf_file_path)