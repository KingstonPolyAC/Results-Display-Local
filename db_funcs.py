import os
import glob
import math
import re
from datetime import datetime
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect
from flask import current_app
from models import WindInfo, TableLog

import sqlite3
from openpyxl import load_workbook
#database, lff_dir, lif_dir = None, None, None

DATABASE = os.path.join(os.getcwd(), "instance", "scoreboard.db") #'instance/scoreboard.db'


def get_last_two_tables():
    logs = TableLog.query.order_by(TableLog.insert_time.desc()).all()
    #Get the last table that has any word Javelin, Discus, Shot, Jump, Pole, Hammer, Throw
    lff_table = next((table for table in logs if any(word in table.table_name for word in ["Javelin", "Discus", "Shot", "Jump", "Pole", "Hammer", "Throw"])), None)
    #Ge the last table that doesnot have any word Javelin, Discus, Shot, Jump, Pole, Hammer, Throw
    lif_table = next((table for table in logs if not any(word in table.table_name for word in ["Javelin", "Discus", "Shot", "Jump", "Pole", "Hammer", "Throw"])), None)
    return (lff_table, lif_table)

def query_last_two_tables(db: SQLAlchemy):
    logs = get_last_two_tables()
    tables_data = {}
    for log in logs:
        table_name = log.table_name
        table_class = db.Model.metadata.tables.get(table_name)
        if table_class is not None:
            rows = db.session.query(table_class).all()
            tables_data[table_name] = rows
    return tables_data

def create_lif_table(db: SQLAlchemy, event_name,heat_no):
    class_name = "".join(word.capitalize() for word in event_name.replace(" ", "_").split("_"))
    
    class TableEvent(db.Model):
        __tablename__ = event_name.replace(" ", "_") #TODO Check for spaces in titles
        __table_args__ = {'extend_existing': True}

        id = db.Column(db.Integer, primary_key=True)
        position = db.Column(db.Integer)
        bib_no = db.Column(db.Integer)
        first_name = db.Column(db.String)
        last_name = db.Column(db.String)
        team = db.Column(db.String)
        time = db.Column(db.String)
        
        

        def __repr__(self):
            return f"<{class_name} {self.first_name} {self.last_name}>"
    #table = TableEvent(event_name, class_name)
     
    # Use inspect to check if the table exists
    inspector = inspect(db.engine)
    if not inspector.has_table(TableEvent.__tablename__):
        db.create_all()  # Create table only if it doesn't exist
        
        # Log the table creation
        table_log = TableLog(table_name=TableEvent.__tablename__,heat_no=heat_no, insert_time=datetime.utcnow())
        db.session.add(table_log)
        db.session.commit()
        
    return TableEvent

def create_lff_table(db: SQLAlchemy, event_name, heat_no):
    class_name = "".join(word.capitalize() for word in event_name.replace(" ", "_").split("_"))
    
    class FieldEvent(db.Model):
        __tablename__ = event_name.replace(" ", "_") #TODO Check for spaces in titles
        __table_args__ = {'extend_existing': True}

        id = db.Column(db.Integer, primary_key=True)
        position = db.Column(db.Integer)
        bib_no = db.Column(db.Integer)
        competition_order = db.Column(db.Integer)
        first_name = db.Column(db.String)
        last_name = db.Column(db.String)
        team = db.Column(db.String)
        performance = db.Column(db.String)
        wind = db.Column(db.String)
        
        

        def __repr__(self):
            return f"<{class_name} {self.first_name} {self.last_name}>"
    #table = TableEvent(event_name, class_name)
     
    # Use inspect to check if the table exists
    inspector = inspect(db.engine)
    if not inspector.has_table(FieldEvent.__tablename__):
        db.create_all()  # Create table only if it doesn't exist
        # Log the table creation
        table_log = TableLog(table_name=FieldEvent.__tablename__, heat_no=heat_no,insert_time=datetime.utcnow())
        db.session.add(table_log)
        db.session.commit()
        
    return FieldEvent

def get_files_sorted_by_mtime(directory):
    # Get a list of files in the directory and filter out directories
    files = [f for f in glob.glob(os.path.join(directory, '*')) if os.path.isfile(f)]
    
    # Get modification time for each file and sort by it
    files_with_mtime = [(file, os.path.getmtime(file)) for file in files]
    sorted_files = sorted(files_with_mtime, key=lambda x: x[1])

    return sorted_files

def convert_to_two_dec_float(value):
    time_string_pattern = re.compile(r'^\d+:\d{2}\.\d{3}$')
    
    if time_string_pattern.match(value):
        string_parts = value.split(".")
        minutes_seconds = string_parts[0]
        milliseconds = string_parts[1]
        
        # Ensure the milliseconds part is exactly three digits
        if len(milliseconds) == 3 and milliseconds.isdigit():
            # Convert milliseconds to float, format to two decimal places
            milliseconds_float = float(milliseconds) / 1000
            formatted_milliseconds = f"{milliseconds_float:.2f}".split('.')[1]
            
            formatted_time = f"{minutes_seconds}.{formatted_milliseconds}"
            return formatted_time
    else:
        try:
            time_value = float(value)
            time_value =  math.ceil(time_value * 100) / 100
            return format(time_value , ".2f")
        except ValueError:
            print("Failed to convert time value: ", value)
            return value

def read_lif_lff_files(db: SQLAlchemy, meeting_info_data):
    """
    Reads LIF and LFF files from the specified directories and updates the database with the data.

    Parameters:
        db (SQLAlchemy): The SQLAlchemy instance representing the database connection.
        lif_directory (str): The directory path where the LIF files are located.
        lff_directory (str): The directory path where the LFF files are located.

    Returns:
        str: The success message if the operation is successful, or an error message if the directories do not exist.
    """

    # global database, lff_dir, lif_dir
    # if db is not None:
    #     database = db
    #     lff_dir = lff_directory
    #     lif_dir = lif_directory
    
    #Check if path exists
    lif_directory = meeting_info_data["track_path"] if meeting_info_data["track_path"] is not None else ""
    lff_directory = meeting_info_data["field_path"] if meeting_info_data["field_path"] is not None else ""
    # if not os.path.exists(lif_directory):
    #     return f"The directory: {lif_directory} does not exist."
    # if not os.path.exists(lff_directory):
    #     return f"The directory: {lff_directory} does not exist."

    print("Processing Track and Field Files to Database")
    #* Check if the track path checkbox has been disabled or 
    if os.path.exists(lif_directory) and meeting_info_data["track_path_check"] == False:
        sorted_files = get_files_sorted_by_mtime(lif_directory)
        #! Reading LIF files
        for filename, mtime in sorted_files:
            if not filename.endswith(".lif"):
                continue
            
            file_path = os.path.join(lif_directory, filename)
            with open(file_path, "r") as file:
                lines = file.readlines()
                event_line = lines[0].strip().split(",")
                event_name = event_line[3] + " " + event_line[2] #Added [2] for Table (A or B) value
                wind_value = event_line[4]+"M/S"
                heat_no = event_line[2]
                print("Event name: ", event_name," Filename: ", filename)
                Event = create_lif_table(db, event_name,heat_no)
                
                inspector = inspect(db.engine)
                if not inspector.has_table("wind_info"):
                    db.create_all()  # Create table only if it doesn't exist
                    #Add data to the wind_info table
                wind_info = WindInfo(table_name=Event.__tablename__, wind_value="0.0M/S" if wind_value == "M/S" else wind_value)
                db.session.add(wind_info)
                db.session.commit()
                
                # if "x" in event_name:
                #     for i, line in enumerate(lines[1:9]):  # Only take up to 8 lines after the first
                #         data = line.strip().split(',')
                        
                #         if data[0].isdigit():
                #             position = int(data[0])
                #             bib_no = int(data[1])
                #             #team = data[2]
                #             first_name = data[4]
                #             last_name = data[3]
                #             team = data[5]
                #             time = data[6]
                # else:
            if len(lines) <= 0:
                continue
            
            for line in lines[1:]:
                data = line.strip().split(",")
                if data[0].isdigit():
                    
                    position = data[0]#int(data[0])
                    bib_no = data[1]#int(data[1])
                    #team = data[2]
                    first_name = data[4]
                    last_name = data[3]
                    team = data[5]
                            
                    time = convert_to_two_dec_float(data[6])
                    print("Time at 6: ", time)
                    #notes = data[6] if len(data) > 6 else ""
                    #wind = data[7] if len(data) > 7 else ""
                    existing_record = db.session.query(Event).filter_by(bib_no=bib_no).first()
                    if existing_record:
                        # Update the existing record if needed
                        existing_record.position = position
                        existing_record.first_name = first_name
                        existing_record.last_name = last_name
                        existing_record.team = team
                        existing_record.time = time
                    else:
                        # Add new record
                        event_instance = Event(position=position, bib_no=bib_no, first_name=first_name,
                                            last_name=last_name, team=team, time=time)
                        db.session.add(event_instance)
                    db.session.commit()
    
    if os.path.exists(lff_directory) and meeting_info_data["field_path_check"] == False:
        sorted_files = get_files_sorted_by_mtime(lff_directory)
        #! Reading LFF Files
        for filename, mtime in sorted_files:
            if not filename.endswith(".lff"):
                continue
            
            file_path = os.path.join(lff_directory, filename)
            with open(file_path, "r") as file:
                lines = file.readlines()
                event_line = lines[0].strip().split(",")
                event_name = event_line[3]
                heat_no = event_line[2]
                print("Event name: ", event_name," Filename: ", filename)
                Event = create_lff_table(db, event_name, heat_no)
                
            if len(lines) < 0:
                continue
            
            for line in lines[1:]:
                data = line.strip().split(",")
                
                position = data[3]#int(data[3])
                bib_no = data[1]
                competition_order = data[2]
                first_name = data[5]
                last_name = data[4]
                team = data[6]
                wind = ""
                
                if "High" not in event_name and "Pole" not in event_name:
                    # Split the input string into a list of performance values
                    def convert_to_two_dec_float(value):
                        try:
                            return format(float(value), ".2f")
                        except ValueError:
                            return value
                        
                    elements = [convert_to_two_dec_float(x) for x in data[7:]]
                    # Initialize variables to keep track of the greatest number and its index
                    greatest_number = float('-inf')  # Start with the smallest possible number
                    greatest_index = -1  # Start with an invalid index
                    
                    # Iterate over the list, checking every other element
                    for i in range(0, len(elements), 2):
                        try:
                            # Convert the element to a float
                            number = float(elements[i])
                            # Check if this number is greater than the current greatest number
                            if number > greatest_number:
                                greatest_number = number
                                greatest_index = i
                        except ValueError:
                            # If conversion fails, ignore this element (though it shouldn't in this context)
                            continue
                    
                    performance = elements[greatest_index] if greatest_index != -1 else ""
                    
                    wind = elements[greatest_index + 1]
                    #notes = data[6] if len(data) > 6 else ""
                    #wind = data[7] if len(data) > 7 else ""
                else:
                    #Get the event heights for High Jump for the event row
                    # ?Find the indices of 'SH' (Start Height) and 'EH' (End Height)
                    start_index = event_line.index('SH') + 1  # Start after 'SH'
                    end_index = event_line.index('EH')  # Stop at 'EH'
                    # ?Extract the parts between 'SH' and 'EH' and convert to float (Will be the event heights)
                    event_heights = [float(event_line[i]) for i in range(start_index, end_index)]
                    
                    # Initialize the variable to store the index of the last occurrence of 'o'
                    last_o_index = -1
                    
                    # Iterate over the parts to find the last 'o'
                    for l_index, part in enumerate(data[7::]):
                        if 'o' in part or 'O' in part:
                            last_o_index = l_index  # Update the index where 'o' is found
                    
                    #* Get the performance value from the event_heights list and format to two decimal places
                    performance = "" if last_o_index == -1 else format(event_heights[last_o_index],".2f")
                    
                    #print("Performance: ", performance)
                #! if the performance value is empty then ignore the entire row
                if performance == "":
                    continue
                
                existing_record = db.session.query(Event).filter_by(bib_no=bib_no).first()
                if existing_record:
                    # Update the existing record if needed
                    existing_record.position = position
                    existing_record.first_name = first_name
                    existing_record.last_name = last_name
                    existing_record.competition_order = competition_order
                    existing_record.team = team
                    existing_record.performance = performance
                    existing_record.wind = wind
                else:
                    # Add new record (include rows that have an athlete name and still register DNS)
                    if first_name != "" and (performance != "DNS" or performance):
                        event_instance = Event(position=position, bib_no=bib_no,competition_order=competition_order, first_name=first_name,
                                            last_name=last_name, team=team, performance=performance, wind=wind)
                        db.session.add(event_instance)
                db.session.commit()
                            
    with current_app.app_context():
        current_app.config['SSE_QUEUE'].put('update')
        
    return "Success"


def insert_data_from_excel_to_db(meeting_info_data):
    excel_path = meeting_info_data["workbook_path"] if meeting_info_data["workbook_path"] is not None else ""
    
    if meeting_info_data["workbook_path_check"]:
        return f"Workbook path ignored"
    #Check if path exists
    if not os.path.exists(excel_path):
        return f"The directory: {excel_path} does not exist."
    
    # Load the workbook and select the sheet
    workbook = load_workbook(filename=excel_path, data_only=True)
    sheet = workbook["MatchDay"]

    # Read data from the specified range
    data = []
    for row in sheet.iter_rows(min_row=23, max_row=29, min_col=2, max_col=4, values_only=True):
        data.append(row)

    # Connect to the SQLite3 database
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    # Create table if it doesn't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS match_results (
            this_match TEXT,
            team_name TEXT,
            points INTEGER
        )
    ''')
    #Delete existing rows
    cursor.execute('''
        DELETE FROM match_results;
    ''')
    # Insert data into the table
    cursor.executemany('''
        INSERT INTO match_results (this_match, team_name, points) 
        VALUES (?, ?, ?)
    ''', data)

    # Commit the changes and close the connection
    conn.commit()
    conn.close()
    
    return "Success"

# excel_path = 'path/to/your/spreadsheet.xlsx'
# db_path = 'path/to/your/database.db'
# insert_data_from_excel_to_db(excel_path)